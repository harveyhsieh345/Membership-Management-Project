import openpyxl as op

class FileOperation:
    def __init__(self, filename):
        # initializing all fields
        self.filename = filename
        self.workbook = op.load_workbook(filename)
        self.sheet = self.workbook.worksheets[0]
        self.headings = []
        self.currentRow = None

        #producing headings
        column = 1
        while True:
            value = self.sheet.cell(row = 1, column = column).value
            if value == None or value == "": break
            self.headings.append(value)
            column += 1

    def search(self, data):
        # If attributes don't match, catch the error without pausing the program
        for i in data:
            if i[0] not in self.headings: 
                self.currentRow = None
                return None

        # Match each row of data to search for the one that fits
        row = 2
        while True:
            if self.sheet.cell(row = row, column = 1).value in ["", None]: 
                break
            flag = True
            for i in data:
                column = self.headings.index(i[0]) + 1
                if str(self.sheet.cell(row = row, column = column).value) != str(i[1]):
                    flag = False
                    break
            if flag:
                self.currentRow = row
                return True
            row += 1

    def returnUserInformation(self, attributes):
        if self.search(attributes) == None: return None

        # Gather all information of the searched user in a list [attribute, value]
        result = []
        for column, heading in enumerate(self.headings):
            value = self.sheet.cell(row = self.currentRow, column = column + 1).value
            result.append([heading, value])
        return result

    def returnAllUserInformation(self):
        # Browse each row and add all information read to a list. The list will contain all users' information.
        row = 2
        result = []
        while True:
            if self.sheet.cell(row = row, column = 1).value in ["", None]: break      
            user = []
            for column, heading in enumerate(self.headings):
                value = self.sheet.cell(row = row, column = column + 1).value
                user.append([heading, value])
            result.append(user)
            row += 1
        return result

    def modifyUserInformation(self, attributes, update):
        if self.search(attributes) == None: return None

        # update each value assigned in list update
        for i in update:
            column = self.headings.index(i[0]) + 1
            self.sheet.cell(row = self.currentRow, column = column).value = i[1]
        self.save()

    def findAllValues(self, singleAttribute):
        # Gathers all existing values for execution.py to prevent duplicate
        if singleAttribute not in self.headings:
            return None
        column = self.headings.index(singleAttribute) + 1
        result = []
        row = 2
        while True:
            if self.sheet.cell(row = row, column = 1).value in [None, ""]: break
            result.append(str(self.sheet.cell(row = row, column = column).value))
            row += 1
        return result

    def addNewMember(self, attributes):
        # Make sure the sequence is correct
        self.updateNumberSequence()

        # Find blank row
        row = 2
        while True:
            if self.sheet.cell(row = row, column = 1).value in ["", None]: break
            row += 1

        # Add all assigned value from the list attributes
        for i in attributes:
            column = self.headings.index(i[0]) + 1
            self.sheet.cell(row = row, column = column).value = i[1]
        
        # update number sequence
        self.updateNumberSequence()

        self.save()

    def deleteMember(self, attributes):
        if self.search(attributes) == None: return None

        # Continue setting current row data to next row data
        while True:
            if self.sheet.cell(row = self.currentRow, column = 1).value in ["", None]: break
            for column, heading in enumerate(self.headings):
                if column == 0: continue
                nextData = self.sheet.cell(row = self.currentRow + 1, column = column + 1).value
                self.sheet.cell(row = self.currentRow, column = column + 1).value = nextData
            self.currentRow += 1
        
        # Readjust number sequence
        self.updateNumberSequence()

        self.save()

    def updateNumberSequence(self):
        row = 2
        while True:
            # Check if the whole row (except column A) is empty
            blank = True
            for column, headings in enumerate(self.headings):
                if column == 0: continue
                if self.sheet.cell(row = row, column = column + 1).value not in ["", None]:
                    blank = False
                    break
            if blank: break

            # Update the number if any cell in the row is not empty
            self.sheet.cell(row = row, column = 1).value = row - 1
            row += 1
        
        # remove unecessary number to prevent error
        for i in range(100):
            self.sheet.cell(row = row, column = 1).value = None
            row += 1

        self.save()

    def checkIfDataIsBlank(self, attributes, singleAttribute):
        # Gather user information
        member = self.returnUserInformation(attributes)
        if singleAttribute not in self.headings or member == None: return None

        # Check if the attribute specified is blank
        index = self.headings.index(singleAttribute)
        if member[index][1] in ["", None]: return True
        return False

    def returnSingleValue(self, attributes, singleAttributes):
        # Gather user information
        member = self.returnUserInformation(attributes)
        if singleAttributes not in self.headings: return None

        # Find the specific value and return
        index = self.headings.index(singleAttributes)
        return member[index][1]

    def save(self):
        # save the file after each edit
        self.workbook.save(self.filename)

